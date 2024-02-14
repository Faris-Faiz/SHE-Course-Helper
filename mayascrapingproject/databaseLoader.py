import json
import os
import openpyxl
from supabase import create_client, Client
from dotenv import load_dotenv

load_dotenv()

supabase_url = os.getenv("SUPABASE_URL")
supabase_key = os.getenv("SUPABASE_KEY")
scraped_data_path = "mayascrapingproject\SHE Course List (df2excel) - 202402141405.xlsx"

supabase: Client = create_client(supabase_url, supabase_key)

def convert_case(string):
    return '_'.join([string.split()[0].lower()] + [word.lower() for word in string.split()[1:]])


def excel_to_json(file_path):
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        headers = [convert_case(cell.value) for cell in sheet[1]]

        data = []

        for index, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True), start=1):
            row_data = dict(zip(headers, row))

            row_data['registered'] = int(row_data['registered'])
            row_data['capacity'] = int(row_data['capacity'])
            row_data['full'] = row_data['full'].lower() == 'true'

            row_data['id'] = index

            data.append(row_data)

        workbook.close()

        with open("scraped_data.json", 'w', encoding='utf-8') as json_file:
            json.dump(data, json_file, ensure_ascii=False, indent=2)

        return data
    except Exception as e:
        print("An exception occurred:", e)
        return []


def on_save_to_supabase(file_path):
    try:
        json_data = excel_to_json(file_path)

        if not json_data:
            print("No data to save to Supabase.")
            return
          
        supabase.table('courses').delete().neq("id", 0).execute()
        supabase.table('courses').upsert(json_data).execute()
        print("!!Uploaded to Supabase Successfully!!")

    except Exception as e:
        print("on_save_to_supabase error:", e)


on_save_to_supabase(scraped_data_path)
